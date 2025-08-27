#!/usr/bin/env python3
"""
Content Hash Validator
Validates that no content is lost during CSV to XLSX conversion process
"""

import csv
import hashlib
import json
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

def extract_csv_content(csv_file_path):
    """
    Extract all question content from original CSV file
    Returns normalized content structure for hashing
    """
    questions = []
    
    with open(csv_file_path, 'r', encoding='utf-8') as file:
        # Skip the first two rows (title and blank row)
        next(file)  # Skip title row
        next(file)  # Skip blank row
        
        reader = csv.DictReader(file)
        
        for row in reader:
            # Skip empty rows or metadata rows
            if not row.get('Question Stem') or not row['Question Stem'].strip():
                continue
                
            question_stem = row['Question Stem'].strip()
            
            # Skip header repetitions, section headers, and metadata
            if (question_stem == 'Question Stem' or 
                'Final Exam' in question_stem or 
                'Quiz' in question_stem or
                not any(row.get(f'Answer {letter}', '').strip() for letter in ['A', 'B', 'C', 'D'])):
                continue
            
            # Extract question data  
            question_number = row.get('Question #', '').strip()
            source_id = f"PREP FL {question_number}" if question_number else "PREP FL"
            
            question_data = {
                'question': question_stem,
                'choices': [
                    row.get('Answer A', '').strip(),
                    row.get('Answer B', '').strip(),
                    row.get('Answer C', '').strip(),
                    row.get('Answer D', '').strip()
                ],
                'correct_answer': row.get('Correct Answer', '').strip(),
                'source': source_id
            }
            
            # Only include questions with actual content
            if question_data['question'] and any(question_data['choices']):
                questions.append(question_data)
    
    return questions

def extract_xlsx_content(xlsx_file_path):
    """
    Extract all question content from converted XLSX file
    Returns normalized content structure for hashing
    """
    questions = []
    
    # Load workbook and get Questions sheet
    wb = load_workbook(xlsx_file_path)
    if 'Questions' not in wb.sheetnames:
        raise ValueError("Questions sheet not found in XLSX file")
    
    ws = wb['Questions']
    
    # Convert to DataFrame for easier processing
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(row)
    
    if not data:
        return questions
    
    # Skip header row
    headers = data[0]
    rows = data[1:]
    
    current_question = None
    current_choices = []
    correct_indices = []
    
    for row in rows:
        if len(row) < 7:
            continue
            
        type_col, question_col, explanation_col, answer_col, correct_col, meta_key_col, meta_value_col = row[:7]
        
        # If we have a question in the first column, it's a new question
        if question_col and str(question_col).strip():
            # Save previous question if exists
            if current_question:
                question_data = {
                    'question': current_question,
                    'choices': current_choices,
                    'correct_answer': convert_indices_to_letter(correct_indices),
                    'source': current_source if 'current_source' in locals() else ''
                }
                questions.append(question_data)
            
            # Start new question
            current_question = str(question_col).strip()
            current_choices = []
            correct_indices = []
            current_source = str(meta_value_col).strip() if meta_value_col else ''
            
            # Add first answer
            if answer_col and str(answer_col).strip():
                current_choices.append(str(answer_col).strip())
                if correct_col == '1' or str(correct_col).strip() == '1':
                    correct_indices.append(0)
        
        # If no question but we have an answer, it's a continuation
        elif answer_col and str(answer_col).strip() and current_question:
            current_choices.append(str(answer_col).strip())
            if correct_col == '1' or str(correct_col).strip() == '1':
                correct_indices.append(len(current_choices) - 1)
    
    # Don't forget the last question
    if current_question:
        question_data = {
            'question': current_question,
            'choices': current_choices,
            'correct_answer': convert_indices_to_letter(correct_indices),
            'source': current_source if 'current_source' in locals() else ''
        }
        questions.append(question_data)
    
    wb.close()
    return questions

def convert_indices_to_letter(indices):
    """Convert list of 0-based indices to letter format (e.g., [3] -> 'D', [0,2] -> 'A, C')"""
    if not indices:
        return ''
    letters = [chr(65 + idx) for idx in sorted(indices) if idx >= 0 and idx < 26]
    return ', '.join(letters) if len(letters) > 1 else letters[0] if letters else ''

def generate_content_hash(questions):
    """
    Generate SHA-256 hash of question content
    Normalizes content to ensure consistent hashing
    """
    # Sort questions by question text for consistent ordering
    sorted_questions = sorted(questions, key=lambda x: x['question'])
    
    # Create normalized content for hashing
    content_for_hash = []
    for q in sorted_questions:
        normalized_q = {
            'question': q['question'].strip().lower(),
            'choices': [choice.strip().lower() for choice in q['choices'] if choice.strip()],
            'correct_answer': q['correct_answer'].strip().upper(),
            'source': q['source'].strip()
        }
        content_for_hash.append(normalized_q)
    
    # Convert to JSON string and hash
    content_json = json.dumps(content_for_hash, sort_keys=True, separators=(',', ':'))
    return hashlib.sha256(content_json.encode('utf-8')).hexdigest()

def validate_conversion(original_csv_path, converted_xlsx_path, output_report_path=None):
    """
    Validate that conversion preserved all content
    Returns validation results and optionally saves detailed report
    """
    print(f"Extracting content from original CSV: {original_csv_path}")
    original_questions = extract_csv_content(original_csv_path)
    original_hash = generate_content_hash(original_questions)
    
    print(f"Extracting content from converted XLSX: {converted_xlsx_path}")
    converted_questions = extract_xlsx_content(converted_xlsx_path)
    converted_hash = generate_content_hash(converted_questions)
    
    # Compare hashes and generate report
    validation_passed = original_hash == converted_hash
    
    report = {
        'validation_passed': validation_passed,
        'original_file': str(original_csv_path),
        'converted_file': str(converted_xlsx_path),
        'original_hash': original_hash,
        'converted_hash': converted_hash,
        'original_question_count': len(original_questions),
        'converted_question_count': len(converted_questions),
        'hash_match': original_hash == converted_hash,
        'count_match': len(original_questions) == len(converted_questions)
    }
    
    # Add detailed comparison if hashes don't match
    if not validation_passed:
        report['differences'] = compare_question_sets(original_questions, converted_questions)
    
    # Save report if path provided
    if output_report_path:
        with open(output_report_path, 'w', encoding='utf-8') as f:
            json.dump(report, f, indent=2, ensure_ascii=False)
        print(f"Detailed report saved to: {output_report_path}")
    
    return report

def compare_question_sets(original, converted):
    """Compare two question sets and identify differences"""
    differences = []
    
    # Create lookup dictionaries
    orig_by_question = {q['question']: q for q in original}
    conv_by_question = {q['question']: q for q in converted}
    
    # Check for missing questions
    missing_in_converted = set(orig_by_question.keys()) - set(conv_by_question.keys())
    extra_in_converted = set(conv_by_question.keys()) - set(orig_by_question.keys())
    
    if missing_in_converted:
        differences.append(f"Questions missing in converted file: {len(missing_in_converted)}")
    
    if extra_in_converted:
        differences.append(f"Extra questions in converted file: {len(extra_in_converted)}")
    
    # Check matching questions for content differences
    common_questions = set(orig_by_question.keys()) & set(conv_by_question.keys())
    for question_text in common_questions:
        orig_q = orig_by_question[question_text]
        conv_q = conv_by_question[question_text]
        
        if orig_q['choices'] != conv_q['choices']:
            differences.append(f"Choice mismatch in question: {question_text[:50]}...")
        
        if orig_q['correct_answer'] != conv_q['correct_answer']:
            differences.append(f"Correct answer mismatch in question: {question_text[:50]}...")
    
    return differences

def main():
    if len(sys.argv) < 3:
        print("Usage: python content_validator.py <original_csv> <converted_xlsx> [report_output_path]")
        sys.exit(1)
    
    original_csv = Path(sys.argv[1])
    converted_xlsx = Path(sys.argv[2])
    report_path = Path(sys.argv[3]) if len(sys.argv) > 3 else None
    
    if not original_csv.exists():
        print(f"Error: Original CSV file {original_csv} not found")
        sys.exit(1)
    
    if not converted_xlsx.exists():
        print(f"Error: Converted XLSX file {converted_xlsx} not found")
        sys.exit(1)
    
    try:
        report = validate_conversion(original_csv, converted_xlsx, report_path)
        
        print("\n" + "="*60)
        print("CONTENT VALIDATION RESULTS")
        print("="*60)
        print(f"Original CSV: {original_csv.name}")
        print(f"Converted XLSX: {converted_xlsx.name}")
        print(f"Original Questions: {report['original_question_count']}")
        print(f"Converted Questions: {report['converted_question_count']}")
        print(f"Question Count Match: {'✅ YES' if report['count_match'] else '❌ NO'}")
        print(f"Content Hash Match: {'✅ YES' if report['hash_match'] else '❌ NO'}")
        print(f"Overall Validation: {'✅ PASSED' if report['validation_passed'] else '❌ FAILED'}")
        
        if not report['validation_passed']:
            print("\nDifferences found:")
            for diff in report.get('differences', []):
                print(f"  - {diff}")
        
        print("="*60)
        
    except Exception as e:
        print(f"❌ Validation error: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Markdown to XLSX Converter - Step 2 of New Approach

Converts standardized markdown format to XLSX with Question Import Template format.
Reuses the excellent XLSX generation and debug logic from the original script.

Usage: python3 markdown_to_xlsx.py input.md output.xlsx
"""

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def parse_markdown_questions(md_path: str) -> tuple[list, dict]:
    """
    Parse standardized markdown and extract questions.
    
    Returns:
        tuple: (list of question dicts, debug stats dict)
    """
    
    with open(md_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # Extract section name
    section_match = re.search(r'^## SECTION: (.+)$', content, re.MULTILINE)
    section_name = section_match.group(1) if section_match else "UNKNOWN-SECTION-SA"
    
    # Split into question blocks
    question_blocks = re.split(r'^### Question \d+', content, flags=re.MULTILINE)[1:]
    
    questions = []
    debug_stats = {
        'tracks_found': [{'track_name': section_name, 'track_number': 1, 'questions_in_track': 0}],
        'total_questions': 0,
        'keep_prefix': False,  # Markdown format doesn't need prefixes
        'parsing_errors': [],
        'detected_pattern': 'Markdown Format',
        'section_format': 'Standardized Markdown',
        'multi_answer_patterns': [],
        'randomization_safe': True  # Will be updated if multi-answer patterns found
    }
    
    # Patterns for multi-answer detection (reused from original script)
    multi_answer_patterns = [
        re.compile(r"\b[A-E]\s+and\s+[A-E](\s+only)?\b", re.IGNORECASE),
        re.compile(r"\b[A-E],\s*[A-E](\s+and\s+[A-E])*(\s+only)?\b", re.IGNORECASE),
        re.compile(r"\b[A-E]\s*,\s*[A-E](\s*,\s*[A-E])*\s*$", re.IGNORECASE),
        re.compile(r"\bBoth\s+[A-E]\s+and\s+[A-E]\b", re.IGNORECASE),
        re.compile(r"\b[A-E]\s+through\s+[A-E]\b", re.IGNORECASE),
        re.compile(r"\b[A-E]\s*[-–]\s*[A-E]\b", re.IGNORECASE),
    ]
    
    question_number = 0
    
    for block in question_blocks:
        if not block.strip():
            continue
            
        question_number += 1
        
        # Extract question text (first line)
        lines = [line.strip() for line in block.split('\n') if line.strip()]
        if not lines:
            continue
            
        question_text = lines[0]
        
        # Extract choices and find correct answer from [X] marker
        choices = []
        correct_index = 1  # Default to first choice if no [X] found
        choice_pattern_checked = re.compile(r'^- \[([xX])\] (.+)$')
        choice_pattern_unchecked = re.compile(r'^- \[ \] (.+)$')
        
        for line in lines:
            # Check for marked correct answer [X] or [x]
            match_checked = choice_pattern_checked.match(line)
            if match_checked:
                choice_text = match_checked.group(2)
                choices.append(choice_text)
                correct_index = len(choices)  # This is the correct answer (1-based index)
                
                # Check for multi-answer patterns
                for pattern in multi_answer_patterns:
                    if pattern.search(choice_text):
                        debug_stats['multi_answer_patterns'].append({
                            'line': question_number,
                            'text': line,
                            'choice_text': choice_text,
                            'pattern': 'Multi-answer detected'
                        })
                        debug_stats['keep_prefix'] = True
                        debug_stats['randomization_safe'] = False
                continue
            
            # Check for unmarked answer [ ]
            match_unchecked = choice_pattern_unchecked.match(line)
            if match_unchecked:
                choice_text = match_unchecked.group(1)
                choices.append(choice_text)
                
                # Check for multi-answer patterns
                for pattern in multi_answer_patterns:
                    if pattern.search(choice_text):
                        debug_stats['multi_answer_patterns'].append({
                            'line': question_number,
                            'text': line,
                            'choice_text': choice_text,
                            'pattern': 'Multi-answer detected'
                        })
                        debug_stats['keep_prefix'] = True
                        debug_stats['randomization_safe'] = False
        
        # Extract metadata
        answer_match = re.search(r'^\*\*Answer:\*\* ([A-E])', block, re.MULTILINE)
        page_match = re.search(r'^\*\*Page:\*\* (.*)$', block, re.MULTILINE)
        section_match = re.search(r'^\*\*Section:\*\* (.*)$', block, re.MULTILINE)
        explanation_match = re.search(r'^\*\*Explanation:\*\* (.*)$', block, re.MULTILINE)
        
        answer_letter = answer_match.group(1) if answer_match else 'A'
        page_ref = page_match.group(1).strip() if page_match else ''
        section_id = section_match.group(1).strip() if section_match else ''
        explanation = explanation_match.group(1).strip() if explanation_match else ''
        
        # NOTE: correct_index is now determined from [X] marker, not Answer letter
        
        # Detect question type based on choices
        question_type = "MC"  # Default to Multiple Choice
        if len(choices) == 2:
            # Check if it's True/False
            choice_texts = [choice.strip().lower() for choice in choices if choice.strip()]
            if set(choice_texts) == {"true", "false"}:
                question_type = "TF"
        
        # Handle choices based on question type
        if question_type == "TF":
            # Keep only 2 choices for True/False
            choices = choices[:2]
        else:
            # For Multiple Choice, keep all choices (up to 6) but pad to minimum of 4
            while len(choices) < 4:
                choices.append("")
            # Don't trim - allow up to 6 choices for MC questions
        
        # Validate question
        if len(question_text) < 5:
            debug_stats['parsing_errors'].append(f"Question {question_number} text too short: '{question_text}'")
            continue
            
        if not any(choices):
            debug_stats['parsing_errors'].append(f"Question {question_number} has no valid choices")
            continue
        
        # Create question object
        question_obj = {
            'track': section_name,
            'question_number': str(question_number),
            'question': question_text,
            'answers': choices,
            'answer_letter': answer_letter,
            'correct': correct_index,
            'explanation': explanation,
            'page_reference': page_ref,
            'section_id': section_id,  # Use section ID as Meta Value
            'question_type': question_type  # Add question type
        }
        
        questions.append(question_obj)
    
    # Update debug stats
    debug_stats['total_questions'] = len(questions)
    debug_stats['tracks_found'][0]['questions_in_track'] = len(questions)
    
    print(f"✅ Parsed {len(questions)} questions from markdown")
    if debug_stats['multi_answer_patterns']:
        print(f"⚠️  Found {len(debug_stats['multi_answer_patterns'])} multi-answer patterns")
    
    return questions, debug_stats

def create_xlsx_output(questions: list, debug_stats: dict, output_path: str):
    """
    Create XLSX file with Question Import Template format plus Debug worksheet.
    (Reused from original script with minor adaptations)
    """
    wb = Workbook()
    
    # === MAIN WORKSHEET (Questions) ===
    ws_main = wb.active
    ws_main.title = "Questions"
    
    # Headers for main worksheet
    headers = ['Type', 'Question', 'Explanation', 'Answer', 'Correct', 'Meta Key', 'Meta Value']
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write questions data
    row = 2
    for q in questions:
        # Row 1 for this question: Main question data
        ws_main.cell(row=row, column=1, value=q.get('question_type', 'MC'))  # Type
        ws_main.cell(row=row, column=2, value=q['question'])  # Question
        ws_main.cell(row=row, column=3, value=q['explanation'])  # Explanation
        ws_main.cell(row=row, column=4, value=q['answers'][0])  # First answer
        if q['correct'] == 1:
            ws_main.cell(row=row, column=5, value='1')  # Correct
        ws_main.cell(row=row, column=6, value='ID')  # Meta Key
        ws_main.cell(row=row, column=7, value=q.get('section_id', q['track']))  # Meta Value (section ID)
        row += 1
        
        # Rows for remaining answer choices
        for answer_idx in range(1, len(q['answers'])):
            if q['answers'][answer_idx]:  # Only write non-empty answers
                ws_main.cell(row=row, column=4, value=q['answers'][answer_idx])
                if q['correct'] == answer_idx + 1:
                    ws_main.cell(row=row, column=5, value='1')
            row += 1
        
        # Ensure correct number of answer rows per question type
        if q.get('question_type') == 'TF':
            # True/False: exactly 2 rows total
            answers_written = len([a for a in q['answers'] if a])
            for _ in range(answers_written, 2):
                row += 1  # Add empty answer row
        else:
            # Multiple Choice: minimum 4 rows, up to actual number of choices
            answers_written = len([a for a in q['answers'] if a])
            min_rows = max(4, answers_written)  # At least 4 rows, more if needed
            for _ in range(answers_written, min_rows):
                row += 1  # Add empty answer row
        
        # Add mandatory blank row between questions
        row += 1
    
    # === DEBUG WORKSHEET ===
    ws_debug = wb.create_sheet("Debug")
    
    # Debug headers
    debug_headers = ['Metric', 'Value']
    for col, header in enumerate(debug_headers, 1):
        cell = ws_debug.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # Debug summary data
    debug_row = 2
    debug_data = [
        ('Total Questions Parsed', debug_stats['total_questions']),
        ('Total Tracks Found', len(debug_stats['tracks_found'])),
        ('Keep Answer Prefixes', 'Yes' if debug_stats['keep_prefix'] else 'No'),
        ('Answer Randomization Safe', 'Yes' if debug_stats['randomization_safe'] else 'No'),
        ('Multi-Answer Patterns Found', len(debug_stats['multi_answer_patterns'])),
        ('Parsing Errors', len(debug_stats['parsing_errors'])),
        ('', ''),  # Blank row
        ('Track Details', ''),
    ]
    
    for metric, value in debug_data:
        ws_debug.cell(row=debug_row, column=1, value=metric)
        ws_debug.cell(row=debug_row, column=2, value=value)
        debug_row += 1
    
    # Track-by-track breakdown
    for track_info in debug_stats['tracks_found']:
        ws_debug.cell(row=debug_row, column=1, value=f"  {track_info['track_name']}")
        ws_debug.cell(row=debug_row, column=2, value=f"{track_info['questions_in_track']} questions")
        debug_row += 1
    
    # Multi-answer patterns found
    if debug_stats['multi_answer_patterns']:
        debug_row += 1
        ws_debug.cell(row=debug_row, column=1, value='Multi-Answer Patterns Found:')
        debug_row += 1
        for pattern_info in debug_stats['multi_answer_patterns']:
            ws_debug.cell(row=debug_row, column=1, value=f"  Question {pattern_info['line']}: {pattern_info['choice_text']}")
            debug_row += 1
    
    # Parsing errors (if any)
    if debug_stats['parsing_errors']:
        debug_row += 1
        ws_debug.cell(row=debug_row, column=1, value='Parsing Errors:')
        debug_row += 1
        for error in debug_stats['parsing_errors']:
            ws_debug.cell(row=debug_row, column=1, value=f"  {error}")
            debug_row += 1
    
    # Question details table
    debug_row += 2
    question_headers = ['Track', 'Q#', 'Question Preview', 'Correct Answer', 'Page Ref', 'Has Explanation']
    for col, header in enumerate(question_headers, 1):
        cell = ws_debug.cell(row=debug_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    debug_row += 1
    
    for q in questions:
        ws_debug.cell(row=debug_row, column=1, value=q['track'])
        ws_debug.cell(row=debug_row, column=2, value=q.get('question_number', '?'))
        ws_debug.cell(row=debug_row, column=3, value=q['question'][:50] + '...' if len(q['question']) > 50 else q['question'])
        ws_debug.cell(row=debug_row, column=4, value=q.get('answer_letter', '?'))
        ws_debug.cell(row=debug_row, column=5, value=q.get('page_reference', ''))
        ws_debug.cell(row=debug_row, column=6, value='Yes' if q['explanation'].strip() else 'No')
        debug_row += 1
    
    # Auto-adjust column widths
    for ws in [ws_main, ws_debug]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 chars
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save workbook
    wb.save(output_path)
    return len(questions)

def main():
    """Main function."""
    if len(sys.argv) < 2:
        print("Usage: python3 markdown_to_xlsx.py input.md [output.xlsx]")
        return 1
    
    input_md = sys.argv[1]
    output_xlsx = sys.argv[2] if len(sys.argv) >= 3 else f"Generated_Output_Files/{Path(input_md).stem}_ready_for_import.xlsx"
    
    if not Path(input_md).exists():
        print(f"Error: Input file '{input_md}' not found.")
        return 1
    
    try:
        # Parse markdown
        questions, debug_stats = parse_markdown_questions(input_md)
        
        if not questions:
            print("Error: No questions found in markdown file.")
            return 1
        
        # Create XLSX
        question_count = create_xlsx_output(questions, debug_stats, output_xlsx)
        
        # Summary
        print(f"✅ Successfully created {output_xlsx}")
        print(f"   - Questions: {question_count}")
        print(f"   - Section: {debug_stats['tracks_found'][0]['track_name']}")
        print(f"   - Answer randomization: {'❌ NOT SAFE' if not debug_stats['randomization_safe'] else '✅ SAFE'}")
        
        return 0
        
    except Exception as e:
        print(f"Error: {e}")
        return 1

if __name__ == "__main__":
    exit(main())